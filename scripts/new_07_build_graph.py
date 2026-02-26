#!/usr/bin/env python3
"""
new_07_build_graph.py
全データを統合して graph-light.json を生成。

入力:
  data/drug_master.json        — 薬マスタ（名寄せ済み）
  data/ddinter_interactions.json — DDI相互作用
  data/cyp_data.json           — CYP代謝情報
  data/adverse_effects_new.json — 副作用
  data/brand_names_new.json    — 商品名
  data/ssk_yakka_master.csv    — SSK薬価基準マスター（薬効分類コード取得）
  /tmp/mhlw_drugs.xlsx         — 厚労省薬価基準Excel（注射薬）
  /tmp/mhlw_usage.xlsx         — 厚労省薬価基準Excel（内用薬）
  data/wikidata_atc.json       — Wikidata DrugBank→ATCマッピング

出力:
  data/graph/graph-light.json  — Cytoscape.js用グラフデータ
"""

import csv
import io
import json
import re
import subprocess
import sys
from pathlib import Path
from collections import Counter, defaultdict

SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR.parent / "data"
GRAPH_DIR = DATA_DIR / "graph"

INPUT_FILES = {
    "drug_master": DATA_DIR / "drug_master.json",
    "ddinter_ddi": DATA_DIR / "ddinter_interactions.json",
    "cyp_data": DATA_DIR / "cyp_data.json",
    "adverse_effects": DATA_DIR / "adverse_effects_new.json",
    "brand_names": DATA_DIR / "brand_names_new.json",
}

SSK_CSV = DATA_DIR / "ssk_yakka_master.csv"
OLD_GRAPH = GRAPH_DIR / "graph-light.json"
MHLW_EXCELS = [Path("/tmp/mhlw_drugs.xlsx"), Path("/tmp/mhlw_usage.xlsx")]
WIKIDATA_ATC = DATA_DIR / "wikidata_atc.json"

OUTPUT = GRAPH_DIR / "graph-light.json"

# ATC level 1-2 → 薬効分類3-4桁コード（最長一致で検索）
ATC_TO_JTC = {
    # A: 消化器官・代謝
    "A02BA": "2323",  # H2遮断剤
    "A02BC": "2329",  # PPI
    "A02B": "232",    # 消化性潰瘍用剤
    "A03": "124",     # 鎮けい剤
    "A04A": "2391",   # 制吐剤
    "A05": "236",     # 利胆剤
    "A06": "235",     # 下剤
    "A07": "231",     # 止しゃ剤・整腸剤
    "A09": "237",     # 消化酵素
    "A10A": "3969",   # インスリン
    "A10B": "396",    # 糖尿病用剤（経口）
    "A11": "319",     # ビタミン剤
    "A12A": "321",    # カルシウム剤
    "A12B": "322",    # 無機質製剤
    "A16": "399",     # その他の代謝性医薬品
    # B: 血液
    "B01AA": "3332",  # ワルファリン類
    "B01AB": "3339",  # ヘパリン類
    "B01AC": "3399",  # 抗血小板剤→その他の血液用薬
    "B01AE": "3339",  # 直接トロンビン阻害
    "B01AF": "3339",  # 直接Xa阻害
    "B01A": "333",    # 血液凝固阻止剤
    "B02": "332",     # 止血剤
    "B03": "322",     # 貧血用剤→無機質製剤
    "B05": "331",     # 血液代用剤
    # C: 循環器
    "C01A": "211",    # 強心配糖体
    "C01B": "212",    # 不整脈用剤
    "C01C": "211",    # 強心剤
    "C01D": "217",    # 血管拡張剤
    "C02": "214",     # 血圧降下剤
    "C03": "213",     # 利尿剤
    "C07": "212",     # β遮断剤
    "C08": "217",     # Ca拮抗剤
    "C09A": "214",    # ACE阻害剤
    "C09B": "214",    # ACE阻害+利尿配合
    "C09C": "214",    # ARB
    "C09D": "214",    # ARB配合
    "C09": "214",     # RAS阻害剤
    "C10AA": "2183",  # HMG-CoA還元酵素阻害
    "C10A": "218",    # 高脂血症用剤
    "C10": "218",     # 高脂血症用剤
    # D: 皮膚科
    "D01": "265",     # 抗真菌→寄生性皮膚疾患用剤
    "D06": "266",     # 化膿性疾患用剤
    "D07": "2452",    # 副腎皮質ステロイド（外用）
    "D10": "269",     # その他の外皮用薬
    # G: 泌尿器・生殖器
    "G01": "252",     # 生殖器官用剤
    "G02": "253",     # 子宮収縮剤
    "G03A": "254",    # 避妊剤
    "G03C": "247",    # 卵胞ホルモン
    "G03D": "247",    # 黄体ホルモン
    "G03H": "246",    # 抗アンドロゲン
    "G04": "251",     # 泌尿器官用剤
    # H: ホルモン
    "H01": "241",     # 脳下垂体ホルモン
    "H02": "245",     # 副腎ホルモン剤
    "H03": "243",     # 甲状腺ホルモン
    "H04": "249",     # その他のホルモン
    "H05": "243",     # 副甲状腺ホルモン
    # J: 抗感染症
    "J01A": "6152",   # テトラサイクリン系
    "J01CA": "6132",  # 広域ペニシリン
    "J01CE": "6131",  # ペニシリン系
    "J01CF": "6131",  # ペニシリン系
    "J01CR": "6132",  # ペニシリン配合
    "J01C": "613",    # ペニシリン系広域
    "J01DB": "6131",  # セフェム系第1世代
    "J01DC": "6132",  # セフェム系第2世代
    "J01DD": "6132",  # セフェム系第3世代
    "J01DE": "6132",  # セフェム系第4世代
    "J01D": "613",    # セフェム系
    "J01FA": "6141",  # マクロライド系
    "J01F": "614",    # マクロライド・リンコマイシン系
    "J01GB": "612",   # アミノグリコシド系
    "J01MA": "6241",  # ニューキノロン系
    "J01M": "624",    # 合成抗菌剤
    "J01XA": "611",   # グリコペプチド系
    "J01": "613",     # 抗菌薬（その他）
    "J02": "617",     # 抗真菌剤
    "J04": "616",     # 抗結核・抗酸菌
    "J05A": "625",    # 抗ウイルス剤（直接作用）
    "J05": "622",     # 抗ウイルス剤
    # L: 抗腫瘍・免疫
    "L01A": "421",    # アルキル化剤
    "L01B": "422",    # 代謝拮抗剤
    "L01C": "424",    # 植物アルカロイド
    "L01D": "423",    # 抗腫瘍性抗生物質
    "L01E": "429",    # 分子標的薬（キナーゼ阻害等）
    "L01F": "429",    # 分子標的薬（抗体）
    "L01X": "429",    # その他の腫瘍用薬
    "L01": "429",     # 抗腫瘍用薬
    "L02": "429",     # ホルモン療法（腫瘍）
    "L03": "449",     # 免疫刺激剤
    "L04": "449",     # 免疫抑制剤
    # M: 筋骨格
    "M01A": "114",    # 解熱鎮痛消炎剤（NSAIDs）
    "M03": "122",     # 骨格筋弛緩剤
    "M04": "394",     # 痛風治療剤
    "M05": "399",     # 骨粗鬆症→代謝性医薬品
    # N: 神経系
    "N01A": "111",    # 全身麻酔剤
    "N01B": "121",    # 局所麻酔剤
    "N02A": "811",    # オピオイド鎮痛
    "N02B": "114",    # 解熱鎮痛消炎剤
    "N02C": "119",    # 片頭痛治療→その他の中枢神経系用薬
    "N03": "113",     # 抗てんかん剤
    "N04": "116",     # 抗パーキンソン剤
    "N05A": "117",    # 抗精神病薬
    "N05B": "112",    # 抗不安剤
    "N05C": "112",    # 催眠鎮静剤
    "N06A": "117",    # 抗うつ剤→精神神経用剤
    "N06B": "115",    # 中枢興奮剤（ADHD等）
    "N06D": "119",    # 抗認知症薬→その他の中枢神経系用薬
    "N07": "119",     # その他の中枢神経系用薬
    # P: 抗寄生虫
    "P01": "625",     # 抗原虫剤
    "P02": "629",     # 駆虫剤→その他の化学療法剤
    # R: 呼吸器
    "R01": "132",     # 耳鼻科用剤（鼻用）
    "R03A": "225",    # β2刺激（吸入）
    "R03B": "225",    # 気管支喘息治療剤（吸入ステロイド等）
    "R03C": "221",    # 気管支拡張剤
    "R03D": "225",    # 気管支喘息治療剤（テオフィリン等）
    "R03": "225",     # 気管支喘息治療剤
    "R05C": "223",    # 去たん剤
    "R05D": "224",    # 鎮咳剤
    "R06": "125",     # 抗ヒスタミン剤
    # S: 感覚器
    "S01": "131",     # 眼科用剤
    "S02": "132",     # 耳鼻科用剤
    # V: その他
    "V03AB": "392",   # 解毒剤
    "V08": "399",     # 造影剤→その他
}

# 日本語の薬効分類（3桁中分類 + 4桁小分類）
# 4桁コードが見つからない場合は先頭3桁で親カテゴリにフォールバック
THERAPEUTIC_CATEGORIES = {
    # === 3桁（中分類） ===
    "111": "全身麻酔剤", "112": "催眠鎮静剤・抗不安剤", "113": "抗てんかん剤",
    "114": "解熱鎮痛消炎剤", "115": "覚せい剤・精神神経用剤", "116": "抗パーキンソン剤",
    "117": "精神神経用剤", "118": "総合感冒剤", "119": "その他の中枢神経系用薬",
    "121": "局所麻酔剤", "122": "骨格筋弛緩剤", "123": "自律神経剤",
    "124": "鎮けい剤", "125": "抗ヒスタミン剤",
    "131": "眼科用剤", "132": "耳鼻科用剤", "133": "鎮暈剤",
    "211": "強心剤", "212": "不整脈用剤", "213": "利尿剤",
    "214": "血圧降下剤", "215": "血管補強剤", "216": "血管収縮剤",
    "217": "血管拡張剤", "218": "高脂血症用剤", "219": "その他の循環器官用薬",
    "221": "気管支拡張剤", "222": "含嗽剤", "223": "去たん剤",
    "224": "鎮咳剤", "225": "気管支喘息治療剤", "226": "呼吸促進剤",
    "229": "その他の呼吸器官用薬",
    "231": "止しゃ剤・整腸剤", "232": "消化性潰瘍用剤", "233": "健胃消化剤",
    "234": "制酸剤", "235": "下剤・浣腸剤", "236": "利胆剤",
    "237": "膵臓疾患用剤", "239": "その他の消化器官用薬",
    "241": "脳下垂体ホルモン剤", "242": "唾液腺ホルモン剤",
    "243": "甲状腺・副甲状腺ホルモン剤", "244": "たん白同化ステロイド剤",
    "245": "副腎ホルモン剤", "246": "男性ホルモン剤", "247": "卵胞・黄体ホルモン剤",
    "248": "混合ホルモン剤", "249": "その他のホルモン剤",
    "251": "泌尿器官用剤", "252": "生殖器官用剤", "253": "子宮収縮剤",
    "254": "避妊剤", "259": "その他の泌尿生殖器官・肛門用薬",
    "261": "外皮用殺菌消毒剤", "264": "鎮痛・鎮痒・収斂・消炎剤",
    "265": "寄生性皮膚疾患用剤", "266": "化膿性疾患用剤",
    "267": "皮膚軟化剤", "269": "その他の外皮用薬",
    "311": "ビタミンA剤", "312": "ビタミンB剤", "313": "ビタミンC剤",
    "314": "ビタミンD剤", "315": "ビタミンE剤", "316": "ビタミンK剤",
    "317": "混合ビタミン剤", "319": "その他のビタミン剤",
    "321": "カルシウム剤", "322": "無機質製剤", "323": "糖類剤",
    "324": "有機酸製剤", "325": "たん白質アミノ酸製剤", "329": "その他の栄養剤",
    "331": "血液代用剤", "332": "止血剤", "333": "血液凝固阻止剤",
    "339": "その他の血液・体液用薬",
    "391": "肝臓疾患用剤", "392": "解毒剤", "393": "習慣性中毒用剤",
    "394": "痛風治療剤", "395": "酵素製剤", "396": "糖尿病用剤",
    "399": "他に分類されない代謝性医薬品",
    "421": "アルキル化剤", "422": "代謝拮抗剤", "423": "抗腫瘍性抗生物質製剤",
    "424": "抗腫瘍性植物成分製剤", "429": "その他の腫瘍用薬",
    "441": "抗アレルギー剤", "442": "刺激療法剤", "449": "その他のアレルギー用薬",
    "611": "主としてグラム陽性菌に作用するもの",
    "612": "主としてグラム陰性菌に作用するもの",
    "613": "主としてグラム陽性・陰性菌に作用するもの",
    "614": "主としてグラム陽性菌・マイコプラズマに作用するもの",
    "615": "主としてグラム陰性菌・リケッチア・クラミジアに作用するもの",
    "616": "主として抗酸菌に作用するもの",
    "617": "主としてカビに作用するもの", "619": "その他の抗生物質製剤",
    "621": "サルファ剤", "622": "抗ウイルス剤", "623": "抗結核剤",
    "624": "合成抗菌剤", "625": "抗原虫剤", "629": "その他の化学療法剤",
    "631": "ワクチン類", "632": "毒素・トキソイド類",
    "634": "血液製剤類", "639": "その他の生物学的製剤",
    "811": "あへんアルカロイド系麻薬", "821": "合成麻薬",
    # === 4桁（小分類）===
    "1115": "バルビツール酸系製剤", "1116": "亜酸化窒素",
    "1119": "その他の全身麻酔剤",
    "1123": "催眠鎮静剤", "1124": "ベンゾジアゼピン系製剤",
    "1125": "バルビツール酸系製剤", "1126": "催眠鎮静剤",
    "1129": "その他の催眠鎮静剤",
    "1131": "フェナセミド系製剤", "1132": "ヒダントイン系製剤",
    "1133": "抗てんかん剤", "1135": "抗てんかん剤",
    "1137": "抗てんかん剤", "1139": "その他の抗てんかん剤",
    "1141": "アニリン系製剤", "1143": "サリチル酸系製剤",
    "1145": "インドメタシン系製剤", "1147": "解熱鎮痛消炎剤",
    "1148": "解熱鎮痛消炎剤", "1149": "その他の解熱鎮痛消炎剤",
    "1151": "中枢興奮剤", "1152": "精神神経用剤",
    "1159": "その他の精神神経用剤",
    "1160": "抗パーキンソン剤", "1169": "その他の抗パーキンソン剤",
    "1171": "フェノチアジン系製剤", "1172": "精神神経用剤",
    "1179": "その他の精神神経用剤",
    "1190": "中枢神経系用薬", "1199": "その他の中枢神経系用薬",
    "1211": "局所麻酔剤", "1214": "局所麻酔剤",
    "1221": "骨格筋弛緩剤", "1229": "その他の骨格筋弛緩剤",
    "1231": "自律神経剤",
    "1241": "鎮けい剤", "1249": "その他の鎮けい剤",
    "1259": "その他の抗ヒスタミン剤",
    "2112": "強心配糖体製剤", "2119": "その他の強心剤",
    "2121": "不整脈用剤", "2123": "β遮断剤", "2129": "その他の不整脈用剤",
    "2131": "チアジド系製剤", "2132": "ループ利尿剤", "2139": "その他の利尿剤",
    "2141": "降圧レセルピン系製剤", "2144": "ACE阻害剤",
    "2149": "その他の血圧降下剤",
    "2171": "カルシウム拮抗剤", "2179": "その他の血管拡張剤",
    "2183": "HMG-CoA還元酵素阻害剤", "2189": "その他の高脂血症用剤",
    "2190": "循環器官用薬",
    "2219": "気管支拡張剤", "2259": "気管支喘息治療剤",
    "2291": "その他の呼吸器官用薬",
    "2319": "止しゃ剤", "2323": "H2遮断剤", "2325": "消化性潰瘍用剤",
    "2329": "プロトンポンプ阻害剤", "2399": "その他の消化器官用薬",
    "2391": "制吐剤",
    "2452": "副腎皮質ステロイド", "2454": "副腎皮質ステロイド",
    "2474": "卵胞ホルモン", "2478": "黄体ホルモン",
    "3332": "クマリン系製剤", "3339": "その他の血液凝固阻止剤",
    "3929": "その他の解毒剤", "3941": "痛風治療剤",
    "3961": "スルホニル尿素系製剤", "3969": "その他の糖尿病用剤",
    "3999": "他に分類されない代謝性医薬品",
    "4211": "アルキル化剤", "4219": "その他のアルキル化剤",
    "4221": "代謝拮抗剤", "4229": "その他の代謝拮抗剤",
    "4231": "抗腫瘍性抗生物質", "4235": "抗腫瘍性抗生物質",
    "4241": "抗腫瘍性植物成分製剤",
    "4291": "その他の腫瘍用薬",
    "4413": "抗アレルギー剤", "4490": "その他のアレルギー用薬",
    "6131": "ペニシリン系", "6132": "広域ペニシリン系",
    "6141": "マクロライド系", "6149": "その他の抗生物質",
    "6152": "テトラサイクリン系",
    "6171": "ポリエン系抗真菌", "6179": "その他の抗真菌剤",
    "6231": "抗結核剤", "6241": "ニューキノロン系",
    "6250": "抗ウイルス剤", "6259": "その他の抗ウイルス剤",
    "6290": "その他の化学療法剤",
    "8114": "モルヒネ系製剤", "8119": "その他のあへんアルカロイド系",
    "8211": "合成麻薬", "8219": "その他の合成麻薬",
}

# 薬効分類推定（英名の接尾辞ベース、4桁コード）
NAME_TO_CATEGORY_REGEX = {
    r"statin$": "2183",      # HMG-CoA還元酵素阻害剤
    r"sartan$": "2149",      # ARB（降圧）
    r"pril$": "2144",        # ACE阻害（降圧）
    r"dipine$": "2171",      # Ca拮抗（降圧）
    r"olol$": "2123",        # β遮断（降圧・不整脈）
    r"semide$": "2132",      # ループ利尿
    r"thiazide$": "2132",    # チアジド利尿
    r"gliptin$": "3969",     # DPP-4阻害（糖尿病）
    r"gliflozin$": "3969",   # SGLT2阻害（糖尿病）
    r"glutide$": "3969",     # GLP-1（糖尿病）
    r"glimepiride|glipizide|glyburide|glibenclamide": "3961",  # SU剤
    r"prazole$": "2329",     # PPI
    r"floxacin$": "6241",    # ニューキノロン
    r"cillin$": "6131",      # ペニシリン系
    r"mycin$": "6141",       # マクロライド系
    r"cycline$": "6152",     # テトラサイクリン系
    r"azole$": "6179",       # 抗真菌
    r"(pam|lam|zepam|zolam)$": "1124",  # ベンゾジアゼピン
    r"barbit(al|one)$": "1125",  # バルビツール酸
    r"(oxetine|aline|amine|pramine)$": "1179",  # 抗うつ
    r"(apine|idone|peridol)$": "1179",    # 抗精神病
    r"profen$": "1149",      # NSAIDs
    r"(fenac|coxib)$": "1149",  # NSAIDs
    r"(xaban|gatran)$": "3339",  # DOAC（抗凝固）
    r"farin$": "3332",       # ワルファリン
    r"(platin)$": "4291",    # 白金系抗がん
    r"(taxel|taxane)$": "4241",  # タキサン系
    r"(rubicin)$": "4231",   # 抗腫瘍性抗生物質
    r"(mab|zumab|ximab|mumab)$": "4291",  # 分子標的薬（抗体）
    r"(tinib|nib)$": "4291",  # 分子標的薬（キナーゼ阻害）
    r"vir$": "6250",         # 抗ウイルス
    r"navir$": "6250",       # プロテアーゼ阻害（HIV）
    r"caine$": "1214",       # 局所麻酔
    r"(solone|sone|olone)$": "2452",  # 副腎皮質ステロイド
    r"(diol|estradiol)$": "2474",  # 卵胞ホルモン
    r"lukast$": "4490",      # ロイコトリエン拮抗
    r"(tadine|tizine)$": "4413",  # 抗ヒスタミン
    r"terol$": "2259",       # β2刺激（気管支拡張）
    r"(setron|ansetron)$": "2391",  # 5-HT3拮抗（制吐）
    r"(done|orphan)$": "8114",  # オピオイド
    r"(pamil|tilazem)$": "2171",  # Ca拮抗
}

# 日本語名先頭一致→薬効分類マッピング（英名がない日本固有薬のフォールバック）
JA_PREFIX_TO_CATEGORY = {
    "インスリン": "3969",        # 糖尿病用剤
    "エポエチン": "339",         # 血液用薬（造血因子）
    "ダルベポエチン": "339",     # 血液用薬（造血因子）
    "フィルグラスチム": "339",   # 血液用薬（G-CSF）
    "エタネルセプト": "449",     # 免疫抑制（生物学的製剤）
    "アダリムマブ": "449",       # 免疫抑制（抗TNFα）
    "インフリキシマブ": "449",   # 免疫抑制（抗TNFα）
    "ウステキヌマブ": "449",     # 免疫抑制（抗IL）
    "トラスツズマブ": "429",     # 抗腫瘍（抗体）
    "ベバシズマブ": "429",       # 抗腫瘍（抗体）
    "リツキシマブ": "429",       # 抗腫瘍（抗体）
    "ニボルマブ": "429",         # 抗腫瘍（免疫CP）
    "ペムブロリズマブ": "429",   # 抗腫瘍（免疫CP）
    "テリパラチド": "243",       # 副甲状腺ホルモン
    "ヨウ化": "243",             # 甲状腺関連
    "ヨード": "243",             # 甲状腺関連
    "ガリウム": "429",           # 放射性医薬品（腫瘍）
    "ヒト(遺伝子組換え": "339",  # 遺伝子組換え血液製剤
    "グリコピロニウム": "225",   # 気管支喘息治療剤（LAMA）
    "ウメクリジニウム": "225",   # 気管支喘息治療剤（LAMA）
    "アクリノール": "261",       # 外皮用殺菌消毒剤
    "テノホビル": "622",         # 抗ウイルス剤
    "エムトリシタビン": "622",   # 抗ウイルス剤
    "ダルナビル": "622",         # 抗ウイルス剤
}

# SSK剤形リスト（名前クリーニング用）
SSK_DOSAGE_FORMS = [
    "ドライシロップ", "シロップ", "カプセル", "ローション",
    "スプレー", "クリーム", "テープ", "パップ", "吸入", "点眼",
    "点鼻", "注射", "軟膏", "顆粒", "細粒", "散", "錠", "液",
    "坐剤", "ゼリー", "ゲル", "粉末",
]


def _extract_ssk_ingredient(generic: str) -> str:
    """【般】ファモチジン散２％ → ファモチジン"""
    name = generic.replace("【般】", "").strip()
    for form in sorted(SSK_DOSAGE_FORMS, key=len, reverse=True):
        idx = name.find(form)
        if idx > 0:
            name = name[:idx]
            break
    name = re.sub(r"[\d０-９．・％%ｍｇμＬｋｇｍＬ]+$", "", name).strip()
    name = re.sub(r"（.*?）$", "", name).strip()
    return name


def _clean_biosimilar(name: str) -> str:
    """バイオシミラー名から後続品表記・括弧を除去
    例: アダリムマブ［アダリムマブ後続１］ → アダリムマブ
    """
    cleaned = re.sub(r"［.*?］", "", name).strip()
    cleaned = re.sub(r"（.*?）", "", cleaned).strip()
    return cleaned


def _strip_salt(name: str) -> str:
    """塩酸塩などの塩形式を除去"""
    return re.sub(
        r"(塩酸塩|硫酸塩|酢酸エステル|酢酸|リン酸|マレイン酸|フマル酸|"
        r"コハク酸|酒石酸|安息香酸|臭化水素酸|メシル酸|ベシル酸|トシル酸|"
        r"ナトリウム|カリウム|カルシウム|マグネシウム|水和物|無水物|エステル)$",
        "", name
    ).strip()


def _atc_to_jtc(atc_code: str) -> str:
    """ATC コードから薬効分類コードへ変換（最長一致）"""
    if not atc_code:
        return ""
    # 最長一致: "J05AG" → "J05A" → "J05" → "J" の順で検索
    for length in range(len(atc_code), 0, -1):
        prefix = atc_code[:length]
        if prefix in ATC_TO_JTC:
            return ATC_TO_JTC[prefix]
    return ""


def _load_mhlw_excel() -> dict:
    """厚労省薬価基準Excelから成分名→4桁薬効分類コード辞書を構築"""
    mhlw_dict = {}
    try:
        import openpyxl
    except ImportError:
        print("  MHLW Excel: openpyxl not installed (skipping)")
        return mhlw_dict

    for xlsx_path in MHLW_EXCELS:
        if not xlsx_path.exists():
            print(f"  MHLW Excel: {xlsx_path.name} not found")
            continue
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # ヘッダー行
            code12 = str(row[1] or "").strip()
            ingredient = str(row[2] or "").strip()
            if len(code12) >= 4 and ingredient:
                code4 = code12[:4]
                mhlw_dict.setdefault(ingredient, code4)
                count += 1
        wb.close()
        print(f"  MHLW Excel: {xlsx_path.name} → {count} rows")

    print(f"  MHLW Excel: {len(mhlw_dict)} unique ingredients, "
          f"{len(set(mhlw_dict.values()))} codes")
    return mhlw_dict


def _load_wikidata_atc() -> dict:
    """Wikidata ATC マッピングを読み込み（DrugBank ID → ATC コードリスト）"""
    if not WIKIDATA_ATC.exists():
        print("  Wikidata ATC: not found (skipping)")
        return {}
    with open(WIKIDATA_ATC, encoding="utf-8") as f:
        data = json.load(f)
    print(f"  Wikidata ATC: {len(data)} DrugBank IDs loaded")
    return data


def build_category_lookup(drugs: list, brand_data: dict) -> dict:
    """複数ソースから薬効分類コード（4桁）を解決する。

    Priority:
      0. 厚労省薬価基準Excel（成分名→4桁コード）— 最も権威的
      1. SSK薬価基準マスター（一般名マッチ）
      2. Wikidata ATC→薬効分類変換
      3. 旧graph-light.json（name_enマッチ）— KEGGベース
      4. 英名接尾辞の正規表現推定
      5. 日本語名先頭一致（バイオ医薬品・インスリン等）
    """
    drug_to_code = {}
    source_stats = Counter()

    # ===== Source 0: 厚労省薬価基準 Excel =====
    mhlw_name_to_code = _load_mhlw_excel()

    # ===== Source 1: SSK 薬価基準マスター =====
    ssk_name_to_code = {}
    if SSK_CSV.exists():
        with open(SSK_CSV, "rb") as f:
            text = f.read().decode("shift_jis", errors="replace")
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            if len(row) < 38:
                continue
            yj = row[31]
            generic = row[37].strip()
            if len(yj) < 4 or not generic.startswith("【般】"):
                continue
            code4 = yj[:4]
            ing = _extract_ssk_ingredient(generic)
            if len(ing) >= 2:
                ssk_name_to_code.setdefault(ing, code4)

        # Also build brand→code for additional matching
        ssk_brand_to_code = {}
        reader2 = csv.reader(io.StringIO(text))
        for row in reader2:
            if len(row) < 35:
                continue
            yj = row[31]
            if len(yj) < 4:
                continue
            brand = row[4].strip()
            for form in sorted(SSK_DOSAGE_FORMS, key=len, reverse=True):
                idx = brand.find(form)
                if idx > 0:
                    brand = brand[:idx]
                    break
            brand = re.sub(r"[\d０-９．・％%ｍｇμＬ]+$", "", brand).strip()
            if len(brand) >= 2:
                ssk_brand_to_code.setdefault(brand, yj[:4])
        print(f"  SSK: {len(ssk_name_to_code)} generic, {len(ssk_brand_to_code)} brand mappings")
    else:
        ssk_brand_to_code = {}
        print("  SSK: not found (skipping)")

    # ===== Source 2: Wikidata ATC =====
    wikidata_atc = _load_wikidata_atc()

    # ===== Source 3: 旧グラフの薬効分類 =====
    old_tc_map = {}
    # Try reading old graph from git
    try:
        result = subprocess.run(
            ["git", "show", "HEAD~1:data/graph/graph-light.json"],
            capture_output=True, text=True, cwd=str(DATA_DIR.parent)
        )
        if result.returncode == 0:
            old_graph = json.loads(result.stdout)
            for n in old_graph["nodes"]:
                if n.get("type") != "drug":
                    continue
                tc = n.get("therapeutic_category", "")
                if not tc:
                    continue
                sn = n.get("search_name", "").strip().lower()
                if sn:
                    old_tc_map[sn] = tc
                en = re.sub(r"\s*\(.*?\)\s*$", "", n.get("name_en", "")).strip().lower()
                if en:
                    old_tc_map[en] = tc
                first = en.split()[0] if en else ""
                if first and len(first) >= 5:
                    old_tc_map.setdefault(first, tc)
            print(f"  Old graph: {len(old_tc_map)} name→tc mappings")
        else:
            print("  Old graph: not available from git")
    except Exception:
        print("  Old graph: git read failed")

    # ===== Resolve for each drug =====
    for drug in drugs:
        did = drug["id"]
        name_ja = drug.get("name_ja", "")
        yakka_name = drug.get("yakka_name", "")
        name_en = drug.get("name_en", "")
        drugbank_id = drug.get("drugbank_id", "")
        if not name_ja:
            continue

        code = None
        src = None

        # Priority 0: MHLW Excel (成分名→4桁コード)
        # バイオシミラー名クリーニング版も候補に追加
        name_variants = [name_ja, yakka_name]
        for n in [name_ja, yakka_name]:
            if n:
                cleaned = _clean_biosimilar(n)
                if cleaned != n and cleaned not in name_variants:
                    name_variants.append(cleaned)
        for name_try in name_variants:
            if not name_try:
                continue
            code = mhlw_name_to_code.get(name_try)
            if code:
                src = "mhlw_exact"
                break
            norm = _strip_salt(name_try)
            if norm != name_try:
                code = mhlw_name_to_code.get(norm)
                if code:
                    src = "mhlw_norm"
                    break
        if not code:
            # MHLW 部分一致（MHLW成分名 ⊂ drug名 or drug名 ⊂ MHLW成分名）
            for name_try in name_variants:
                if not name_try or len(name_try) < 3:
                    continue
                for ing, c in mhlw_name_to_code.items():
                    if len(ing) >= 3 and (ing in name_try or name_try in ing):
                        code = c
                        src = "mhlw_sub"
                        break
                if code:
                    break

        # Priority 1: SSK generic name
        if not code:
            for name_try in [name_ja, yakka_name]:
                if not name_try:
                    continue
                code = ssk_name_to_code.get(name_try)
                if code:
                    src = "ssk_generic"
                    break
                norm = _strip_salt(name_try)
                code = ssk_name_to_code.get(norm)
                if code:
                    src = "ssk_norm"
                    break
                for ing, c in ssk_name_to_code.items():
                    if len(ing) >= 3 and len(name_try) >= 3:
                        if ing in name_try or name_try in ing:
                            code = c
                            src = "ssk_sub"
                            break
                if code:
                    break

        # Priority 1b: SSK brand name
        if not code:
            drug_brands = brand_data.get(did, [])
            for brand in drug_brands:
                code = ssk_brand_to_code.get(brand)
                if code:
                    src = "ssk_brand"
                    break

        # Priority 2: Wikidata ATC → 薬効分類変換
        if not code and drugbank_id and drugbank_id in wikidata_atc:
            atc_list = wikidata_atc[drugbank_id]
            for atc in atc_list:
                jtc = _atc_to_jtc(atc)
                if jtc:
                    code = jtc
                    src = "wikidata_atc"
                    break

        # Priority 3: Old graph
        if not code and name_en:
            en_lower = name_en.strip().lower()
            code = old_tc_map.get(en_lower)
            if code:
                src = "old_graph"
            else:
                en_base = re.sub(
                    r"\s+(hydrochloride|sodium|potassium|calcium|sulfate|"
                    r"acetate|phosphate|mesylate|besylate|maleate|fumarate|"
                    r"tartrate|succinate|citrate|bromide|chloride)$",
                    "", en_lower
                )
                code = old_tc_map.get(en_base)
                if code:
                    src = "old_norm"
                else:
                    first = en_lower.split()[0]
                    if len(first) >= 5:
                        code = old_tc_map.get(first)
                        if code:
                            src = "old_first"

        # Priority 4: Regex suffix
        if not code and name_en:
            lower = name_en.lower()
            for pattern, c in NAME_TO_CATEGORY_REGEX.items():
                if re.search(pattern, lower):
                    code = c
                    src = "regex"
                    break

        # Priority 5: 日本語名先頭一致（バイオ医薬品・インスリン等）
        if not code and name_ja:
            base = _clean_biosimilar(name_ja)
            for prefix, c in JA_PREFIX_TO_CATEGORY.items():
                if base.startswith(prefix):
                    code = c
                    src = "ja_prefix"
                    break

        if code:
            drug_to_code[did] = code
            source_stats[src] += 1

    total_ja = len([d for d in drugs if d.get("name_ja")])
    print(f"\n  薬効分類解決: {len(drug_to_code)}/{total_ja} "
          f"({len(drug_to_code)/total_ja*100:.1f}%), "
          f"{len(set(drug_to_code.values()))} unique codes")
    for src, cnt in source_stats.most_common():
        print(f"    {src}: {cnt}")

    return drug_to_code


def estimate_category(name_en: str) -> str:
    """英名から薬効分類コードを推定（フォールバック）"""
    lower = name_en.lower()
    for pattern, code in NAME_TO_CATEGORY_REGEX.items():
        if re.search(pattern, lower):
            return code
    return ""


def load_data() -> dict:
    """全入力データを読み込み"""
    data = {}
    for key, path in INPUT_FILES.items():
        if path.exists():
            with open(path, encoding="utf-8") as f:
                data[key] = json.load(f)
            print(f"  {key}: loaded ({path.name})")
        else:
            print(f"  WARNING: {path.name} not found")
            data[key] = {}
    return data


def build_graph(data: dict) -> dict:
    """全データを統合してグラフ構造を構築"""
    master = data.get("drug_master", {})
    all_drugs = master.get("drugs", [])
    ddi_data = data.get("ddinter_ddi", {})
    ddis = ddi_data.get("interactions", [])
    cyp_data = data.get("cyp_data", {}).get("cyp_data", {})
    ae_data = data.get("adverse_effects", {}).get("adverse_effects", [])
    brand_data = data.get("brand_names", {}).get("brand_names", {})

    # ============ 薬効分類コード解決（複数ソース） ============
    print(f"\n=== 薬効分類コード解決 ===")
    drug_to_tc = build_category_lookup(all_drugs, brand_data)

    print(f"\n=== 入力データ（フィルタ前） ===")
    print(f"薬数: {len(all_drugs)}")
    print(f"DDIペア: {len(ddis)}")

    # ============ フィルタ: 日本市場関連薬に絞り込み ============
    # Step 1: name_ja ありの薬を「コア薬」とする
    # Step 2: コア薬のDDIパートナーを追加（name_jaなしでも）
    # Step 3: Minor DDI (Level 1) は除外してサイズ削減
    ddinter_to_id_all = {}
    for d in all_drugs:
        if d.get("ddinter_id"):
            ddinter_to_id_all[d["ddinter_id"]] = d["id"]

    drug_by_id_all = {d["id"]: d for d in all_drugs}

    # コア薬 = name_jaあり
    core_ids = set()
    for d in all_drugs:
        if d.get("name_ja"):
            core_ids.add(d["id"])

    print(f"コア薬 (name_jaあり): {len(core_ids)}")

    # DDIフィルタ: コア薬(name_jaあり)同士のDDIのみ保持
    # さらに Minor DDI (Level 1) は除外
    filtered_ddis = []
    ddi_seen = set()
    for ix in ddis:
        level_str = str(ix.get("level", "")).strip()
        if level_str == "1":  # Minor DDI は除外
            continue
        id_a = ddinter_to_id_all.get(ix["drug_a"], ix["drug_a"])
        id_b = ddinter_to_id_all.get(ix["drug_b"], ix["drug_b"])
        # 両端がコア薬（name_jaあり）
        if id_a in core_ids and id_b in core_ids:
            pair = tuple(sorted([id_a, id_b]))
            if pair not in ddi_seen:
                ddi_seen.add(pair)
                filtered_ddis.append((ix, id_a, id_b, level_str))
    ddis_to_process = filtered_ddis

    # 最終薬リスト = コア薬のみ（DDIパートナーは必然的にコア内）
    drugs = [d for d in all_drugs if d["id"] in core_ids]

    print(f"フィルタ後薬数: {len(drugs)} (除外: {len(all_drugs) - len(drugs)})")
    print(f"DDI (コア薬間, Minor除外, dedup): {len(ddis_to_process)} (元: {len(ddis)})")
    print(f"CYPデータ薬数: {len(cyp_data)}")
    print(f"副作用データ薬数: {len(ae_data)}")
    print(f"商品名マッチ薬数: {len(brand_data)}")

    # Build lookup maps
    drug_by_id = {d["id"]: d for d in drugs}
    # DDinter ID → drug ID
    ddinter_to_id = {}
    for d in drugs:
        if d.get("ddinter_id"):
            ddinter_to_id[d["ddinter_id"]] = d["id"]
    # EN name (lower) → drug ID
    en_to_id = {}
    for d in drugs:
        if d.get("name_en"):
            en_to_id[d["name_en"].lower()] = d["id"]

    ae_by_drug = {ae["drug_id"]: ae for ae in ae_data}

    nodes = []
    edges = []
    category_nodes = {}
    cyp_nodes = {}
    adverse_nodes = {}
    edge_id = 0

    # ============ Drug Nodes ============
    for drug in drugs:
        drug_id = drug["id"]
        name_en = drug.get("name_en", "")
        name_ja = drug.get("name_ja", "")

        # CYP情報
        cyp_enzymes = cyp_data.get(name_en.lower(), [])

        # 副作用
        ae_entry = ae_by_drug.get(drug_id, {})
        adverse_effects = ae_entry.get("adverse_effects", [])

        # 商品名
        brands = brand_data.get(drug_id, [])
        names_alt = brands.copy()

        # 薬効分類（複数ソースから解決済み）
        tc = drug_to_tc.get(drug_id, "")

        node = {
            "id": drug_id,
            "type": "drug",
            "name_en": name_en,
            "name_ja": name_ja,
            "names_alt": names_alt,
            "search_name": name_en,
            "therapeutic_category": tc,
            "cyp_enzymes": cyp_enzymes,
            "adverse_effects": adverse_effects,
            "drugbank_id": drug.get("drugbank_id", ""),
        }
        nodes.append(node)

        # Category node
        if tc and tc not in category_nodes:
            # 4桁→3桁フォールバックで名前解決
            cat_name = THERAPEUTIC_CATEGORIES.get(tc)
            if not cat_name and len(tc) == 4:
                cat_name = THERAPEUTIC_CATEGORIES.get(tc[:3])
            if not cat_name:
                cat_name = f"分類{tc}"
            cat_id = f"cat_{tc}"
            category_nodes[tc] = cat_id
            nodes.append({
                "id": cat_id,
                "type": "category",
                "name_ja": cat_name,
                "name_en": "",
                "code": tc,
            })

        # CYP nodes
        for cyp in cyp_enzymes:
            if cyp not in cyp_nodes:
                cyp_id = f"cyp_{cyp}"
                cyp_nodes[cyp] = cyp_id
                nodes.append({
                    "id": cyp_id,
                    "type": "cyp",
                    "name_en": cyp,
                    "name_ja": cyp,
                })

        # Adverse effect nodes
        for ae in adverse_effects:
            ae_name = ae["name"]
            if ae_name not in adverse_nodes:
                ae_id = f"ae_{len(adverse_nodes)}"
                adverse_nodes[ae_name] = ae_id
                nodes.append({
                    "id": ae_id,
                    "type": "adverse_effect",
                    "name_ja": ae_name,
                    "name_en": ae.get("name_en", ""),
                })

    # ============ Edges ============

    # DDI edges (already filtered in preprocessing)
    valid_ids = {n["id"] for n in nodes if n["type"] == "drug"}
    ddi_stats = Counter()
    seen_ddi_pairs = set()

    for ix_tuple in ddis_to_process:
        ix, id_a, id_b, level_str = ix_tuple

        if id_a not in valid_ids or id_b not in valid_ids:
            ddi_stats["skip_missing"] += 1
            continue

        if id_a == id_b:
            ddi_stats["skip_selfloop"] += 1
            continue

        # Dedup
        pair_key = tuple(sorted([id_a, id_b]))
        if pair_key in seen_ddi_pairs:
            continue
        seen_ddi_pairs.add(pair_key)

        if level_str in ("3", "Major"):
            edge_type = "contraindication"
            severity = "CI"
        else:
            edge_type = "precaution"
            severity = "P"

        edge_id += 1
        edges.append({
            "id": f"e_{edge_id}",
            "source": id_a,
            "target": id_b,
            "type": edge_type,
            "severity": severity,
        })
        ddi_stats[edge_type] += 1

    print(f"\nDDI処理: {dict(ddi_stats)}")

    # Category edges
    for drug in drugs:
        drug_id = drug["id"]
        tc = drug_to_tc.get(drug_id, "")
        if tc and tc in category_nodes:
            edge_id += 1
            edges.append({
                "id": f"e_{edge_id}",
                "source": drug_id,
                "target": category_nodes[tc],
                "type": "belongs_to_category",
            })

    # CYP edges
    for drug in drugs:
        drug_id = drug["id"]
        name_en = drug.get("name_en", "").lower()
        for cyp in cyp_data.get(name_en, []):
            if cyp in cyp_nodes:
                edge_id += 1
                edges.append({
                    "id": f"e_{edge_id}",
                    "source": drug_id,
                    "target": cyp_nodes[cyp],
                    "type": "metabolized_by",
                })

    # Adverse effect edges
    for ae_entry in ae_data:
        drug_id = ae_entry["drug_id"]
        if drug_id not in valid_ids:
            continue
        for ae in ae_entry.get("adverse_effects", []):
            ae_name = ae["name"]
            if ae_name in adverse_nodes:
                edge_id += 1
                edges.append({
                    "id": f"e_{edge_id}",
                    "source": drug_id,
                    "target": adverse_nodes[ae_name],
                    "type": "causes_adverse_effect",
                    "frequency": ae.get("frequency", ""),
                })

    return {"nodes": nodes, "edges": edges}


def validate(graph: dict) -> bool:
    """グラフのバリデーション"""
    print(f"\n=== バリデーション ===")
    nodes = graph["nodes"]
    edges = graph["edges"]
    ok = True

    # 1. ノードID一意性
    ids = [n["id"] for n in nodes]
    dupes = [id for id, cnt in Counter(ids).items() if cnt > 1]
    if dupes:
        print(f"  FAIL: 重複ノードID: {dupes[:5]}")
        ok = False
    else:
        print(f"  OK: ノードID一意 ({len(ids)})")

    # 2. エッジ参照整合性
    node_ids = set(ids)
    bad_refs = 0
    for e in edges:
        if e.get("source") not in node_ids:
            bad_refs += 1
        if e.get("target") not in node_ids:
            bad_refs += 1
    if bad_refs:
        print(f"  FAIL: 参照エラー {bad_refs} エッジ")
        ok = False
    else:
        print(f"  OK: エッジ参照整合 ({len(edges)})")

    # 3. 自己ループ
    self_loops = sum(1 for e in edges if e.get("source") == e.get("target"))
    if self_loops:
        print(f"  WARN: 自己ループ {self_loops} 件")
    else:
        print(f"  OK: 自己ループなし")

    # 4. DDIがdrug-drug間のみ
    drug_ids = {n["id"] for n in nodes if n["type"] == "drug"}
    ddi_edges = [e for e in edges if e["type"] in ("contraindication", "precaution")]
    bad_ddi = sum(1 for e in ddi_edges
                  if e["source"] not in drug_ids or e["target"] not in drug_ids)
    if bad_ddi:
        print(f"  FAIL: 非drug-drug DDI {bad_ddi} 件")
        ok = False
    else:
        print(f"  OK: DDI全てdrug-drug間")

    return ok


def print_stats(graph: dict):
    """統計を出力"""
    nodes = graph["nodes"]
    edges = graph["edges"]

    nt = Counter(n["type"] for n in nodes)
    et = Counter(e["type"] for e in edges)

    drugs = [n for n in nodes if n["type"] == "drug"]
    has_ja = sum(1 for d in drugs if d.get("name_ja"))
    has_cyp = sum(1 for d in drugs if d.get("cyp_enzymes"))
    has_ae = sum(1 for d in drugs if d.get("adverse_effects"))
    has_alt = sum(1 for d in drugs if d.get("names_alt"))

    print(f"\n=== ノード統計 ===")
    for t, c in nt.most_common():
        print(f"  {t}: {c}")
    print(f"  合計: {len(nodes)}")

    print(f"\n=== エッジ統計 ===")
    for t, c in et.most_common():
        print(f"  {t}: {c}")
    print(f"  合計: {len(edges)}")

    print(f"\n=== カバレッジ ===")
    print(f"  name_ja: {has_ja}/{len(drugs)} ({has_ja/len(drugs)*100:.1f}%)")
    print(f"  CYP: {has_cyp}/{len(drugs)}")
    print(f"  副作用: {has_ae}/{len(drugs)}")
    print(f"  商品名: {has_alt}/{len(drugs)}")

    # 定量比較
    print(f"\n=== 定量比較（現行vs新） ===")
    print(f"  薬ノード: {nt.get('drug', 0)} (現行2,600, 許容2,000-3,000)")
    ci = et.get("contraindication", 0)
    p = et.get("precaution", 0)
    print(f"  DDI(CI): {ci} (現行4,019, 許容2,000-6,000)")
    print(f"  DDI(P): {p} (現行18,327, 許容10,000-30,000)")
    ja_rate = has_ja / len(drugs) * 100 if drugs else 0
    print(f"  name_ja率: {ja_rate:.1f}% (現行79%, 目標>80%)")

    # 主要薬チェック
    print(f"\n=== 主要薬チェック ===")
    check_drugs = [
        "Warfarin", "Loxoprofen", "Clarithromycin", "Amiodarone",
        "Amlodipine", "Lansoprazole", "Carbamazepine", "Cyclosporine",
        "Metformin", "Prednisolone",
    ]
    for name in check_drugs:
        found = [d for d in drugs if d.get("name_en", "").lower() == name.lower()
                 or d.get("search_name", "").lower() == name.lower()]
        if found:
            d = found[0]
            print(f"  {name}: id={d['id']}, ja={d.get('name_ja','?')}, "
                  f"cyp={len(d.get('cyp_enzymes',[]))}, ae={len(d.get('adverse_effects',[]))}")
        else:
            print(f"  {name}: NOT FOUND")


def main():
    print("=== graph-light.json 統合ビルド ===\n")

    # Load
    print("データ読み込み:")
    data = load_data()

    # Build
    graph = build_graph(data)

    # Validate
    valid = validate(graph)

    # Stats
    print_stats(graph)

    # Save
    GRAPH_DIR.mkdir(exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(graph, f, ensure_ascii=False, separators=(",", ":"))

    size_mb = OUTPUT.stat().st_size / 1024 / 1024
    print(f"\n保存: {OUTPUT} ({size_mb:.1f} MB)")

    if not valid:
        print("\nWARNING: バリデーションエラーあり。確認してください。")
        sys.exit(1)
    else:
        print("\nSUCCESS: バリデーション通過")


if __name__ == "__main__":
    main()
